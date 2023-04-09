import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

archivo_excel = pd.read_excel('supermarket_sales.xlsx')

# print(archivo_excel[['Gender', 'Product line', 'Total']])


tabla_pivote = archivo_excel.pivot_table(
    index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
# print(tabla_pivote)

tabla_pivote.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')


wb = load_workbook('sales_2021.xlsx')

pestaña = wb['Report']


# Se identifican las columnas y filas activas del archivo excel

min_col = wb.active.min_column
min_col = wb.active.max_column
min_fila = wb.active.min_row
min_fila = wb.active.max_row

print(min_col)
