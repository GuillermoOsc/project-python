import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

archivo_excel = pd.read_excel('supermarket_sales.xlsx')

# print(archivo_excel[['Gender', 'Product line', 'Total']])


tabla_pivote = archivo_excel.pivot_table(
    index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)


tabla_pivote.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')


wb = load_workbook('sales_2021.xlsx')

pestaña = wb['Report']


# Se identifican las columnas y filas activas del archivo excel

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row


# Graficos

barchart = BarChart()

data = Reference(pestaña, min_col=min_col+1, max_col=max_col,
                 min_row=min_fila, max_row=max_fila)

categorias = Reference(pestaña, min_col=min_col+1, max_col=min_col,
                       min_row=min_fila+1, max_row=max_fila)


barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categorias)

pestaña.add_chart(barchart, 'B12')
barchart.title = 'Ventas'
barchart.style = 2

pestaña['B8'] = '=SUM(B6:B7)'  # Script para aplicar formula de excel.
pestaña['B8'].style = 'Currency'

wb.save('sales_2021.xlsx')
